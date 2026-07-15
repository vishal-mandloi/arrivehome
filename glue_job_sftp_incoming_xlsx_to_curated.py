"""
AWS Glue ETL Job: SFTP Incoming XLSX folders → Curated Parquet (Athena)

Source: one latest.xlsx per vendor folder under S3 incoming prefix.
Example layout:
  s3://arrive-home-sftp-files/incoming/Ext_AbsoluteResolutionsInvestmentsLLC/latest.xlsx
  s3://arrive-home-sftp-files/incoming/Ext_DyckONeal/latest.xlsx
  ...

Each workbook has a single tab. The job reads the first sheet and loads it into
its own Athena table in the arrive_home database (snapshot mode by default).

Required Glue job parameter:
  --additional-python-modules openpyxl

Parameters:
  --SOURCE_S3_PREFIX   (default: s3://arrive-home-sftp-files/incoming/)
  --XLSX_FILENAME      (default: latest.xlsx)
  --FOLDER_TABLE_MAP   (optional JSON: {"FolderName": "athena_table_name", ...})
  --OUTPUT_S3_BASE     (default: s3://arrivehome-bi-prod/curated/)
  --GLUE_DATABASE      (default: arrive_home)
  --WRITE_MODE         (snapshot | daily_partition; default: snapshot)
  --DT                 (default: today UTC)
"""

import io
import json
import re
import sys
from datetime import datetime, timezone
from typing import Dict, List, Tuple

import boto3
from awsglue.context import GlueContext
from awsglue.job import Job
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql.types import StringType, StructField, StructType


def _arg_or_default(flag: str, default: str) -> str:
    if flag in sys.argv:
        return getResolvedOptions(sys.argv, [flag.lstrip("-")])[flag.lstrip("-")]
    return default


_word_re = re.compile(r"[A-Za-z0-9]+")
_camel_split_re = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")


def _to_snake_case(col: str) -> str:
    raw = (col or "").strip()
    if not raw:
        return "col"
    tokens = _word_re.findall(raw)
    words: List[str] = []
    for t in tokens:
        words.extend(p for p in _camel_split_re.split(t) if p)
    name = "_".join(w.lower() for w in words)
    if name and name[0].isdigit():
        name = f"col_{name}"
    return name or "col"


def _folder_to_table_name(folder_name: str) -> str:
    """Ext_AbsoluteResolutionsInvestmentsLLC -> ext_absolute_resolutions_investments_llc"""
    return _to_snake_case(folder_name)


def _dedupe_columns(cols) -> Dict[str, str]:
    seen: Dict[str, int] = {}
    mapping: Dict[str, str] = {}
    for col in cols:
        base = _to_snake_case(col)
        if base not in seen:
            seen[base] = 0
            mapping[col] = base
        else:
            seen[base] += 1
            mapping[col] = f"{base}_{seen[base]}"
    return mapping


def _parse_s3_uri(s3_uri: str) -> Tuple[str, str]:
    if not s3_uri.startswith("s3://"):
        raise ValueError(f"Expected s3:// URI, got: {s3_uri}")
    without = s3_uri[len("s3://") :]
    bucket, _, key = without.partition("/")
    if not bucket:
        raise ValueError(f"Expected s3://bucket/... URI, got: {s3_uri}")
    return bucket, key.lstrip("/")


def _is_numeric(v: str) -> bool:
    if not v:
        return False
    s = v.strip().replace(",", "").replace("$", "").replace("%", "")
    if s.startswith("-"):
        s = s[1:]
    s = s.replace(".", "", 1)
    return s.isdigit()


def _load_xlsx_bytes_from_s3(s3_uri: str) -> bytes:
    bucket, key = _parse_s3_uri(s3_uri)
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    return obj["Body"].read()


def _delete_s3_prefix(s3_uri: str) -> int:
    bucket, prefix = _parse_s3_uri(s3_uri)
    if not prefix.endswith("/"):
        prefix = prefix + "/"
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    deleted = 0
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        objs = [{"Key": o["Key"]} for o in page.get("Contents", [])]
        if not objs:
            continue
        for i in range(0, len(objs), 1000):
            s3.delete_objects(Bucket=bucket, Delete={"Objects": objs[i : i + 1000]})
        deleted += len(objs)
    return deleted


def _open_workbook(xlsx_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required. Add `--additional-python-modules openpyxl` "
            "to the Glue job parameters."
        ) from e
    return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)


def _read_sheet_rows(workbook, sheet_name: str) -> List[List[object]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
        )
    sheet = workbook[sheet_name]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _detect_header_row(all_rows: List[List[object]]) -> int:
    if not all_rows:
        return 0
    max_cols = max(len(r) for r in all_rows)
    for i, row in enumerate(all_rows[:30]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for c in cells if c)
        if non_empty < max(3, max_cols // 2):
            continue
        text_cells = sum(1 for c in cells if c and not _is_numeric(c))
        if text_cells >= non_empty * 0.7:
            return i
    return 0


def _extract_header_and_data(
    all_rows: List[List[object]],
) -> Tuple[List[str], List[Tuple[str, ...]]]:
    if not all_rows:
        return [], []

    header_idx = _detect_header_row(all_rows)
    raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    last_real_col = -1
    for i, h in enumerate(raw_headers):
        if h:
            last_real_col = i
    if last_real_col < 0:
        last_real_col = len(raw_headers) - 1

    width = last_real_col + 1
    headers = raw_headers[:width]
    data = all_rows[header_idx + 1 :]

    print(f"  Detected header row at xlsx row index: {header_idx} (0-based)")
    print(f"  Header reported width: {len(raw_headers)}; trimmed to: {width}")

    norm_rows: List[Tuple[str, ...]] = []
    for r in data:
        cells = [("" if c is None else str(c)) for c in r[:width]]
        if len(cells) < width:
            cells = cells + [""] * (width - len(cells))
        if any(c.strip() for c in cells):
            norm_rows.append(tuple(cells))

    return headers, norm_rows


def _build_spark_df(spark, headers: List[str], data_rows: List[Tuple[str, ...]]) -> DataFrame:
    safe_headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(headers):
        base = h if h else f"col_{i}"
        if base not in seen:
            seen[base] = 0
            safe_headers.append(base)
        else:
            seen[base] += 1
            safe_headers.append(f"{base}_{seen[base]}")

    schema = StructType([StructField(h, StringType(), True) for h in safe_headers])
    return spark.createDataFrame(data_rows, schema=schema)


def _spark_type_to_hive(dtype) -> str:
    return dtype.simpleString()


def _register_glue_table(
    database: str,
    table: str,
    location: str,
    schema_columns: List[Tuple[str, str]],
    partitioned: bool = False,
) -> None:
    glue = boto3.client("glue")

    storage_descriptor = {
        "Columns": [{"Name": name, "Type": dtype} for name, dtype in schema_columns],
        "Location": location.rstrip("/") + "/",
        "InputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetInputFormat",
        "OutputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetOutputFormat",
        "Compressed": True,
        "SerdeInfo": {
            "SerializationLibrary": "org.apache.hadoop.hive.ql.io.parquet.serde.ParquetHiveSerDe",
            "Parameters": {"serialization.format": "1"},
        },
        "StoredAsSubDirectories": False,
    }

    parameters = {"classification": "parquet", "EXTERNAL": "TRUE"}
    partition_keys: List[Dict[str, str]] = []
    if partitioned:
        parameters.update(
            {
                "projection.enabled": "true",
                "projection.dt.type": "date",
                "projection.dt.format": "yyyy-MM-dd",
                "projection.dt.range": "2026-01-01,NOW",
                "projection.dt.interval": "1",
                "projection.dt.interval.unit": "DAYS",
                "storage.location.template": location.rstrip("/") + "/dt=${dt}",
            }
        )
        partition_keys = [{"Name": "dt", "Type": "string"}]

    table_input = {
        "Name": table,
        "TableType": "EXTERNAL_TABLE",
        "Parameters": parameters,
        "StorageDescriptor": storage_descriptor,
        "PartitionKeys": partition_keys,
    }

    try:
        glue.get_database(Name=database)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue database: {database}")
        glue.create_database(DatabaseInput={"Name": database})

    try:
        glue.get_table(DatabaseName=database, Name=table)
        print(f"  Updating Glue table: {database}.{table}")
        glue.update_table(DatabaseName=database, TableInput=table_input)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue table: {database}.{table}")
        glue.create_table(DatabaseName=database, TableInput=table_input)


def _process_and_load(
    spark,
    workbook,
    source_s3_uri: str,
    sheet_name: str,
    table_name: str,
    folder_name: str,
    glue_database: str,
    output_s3_base: str,
    write_mode: str,
    dt: str,
    loaded_at: str,
) -> None:
    print(f"  Sheet names in workbook: {workbook.sheetnames}")

    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found in {source_s3_uri}. "
            f"Available: {workbook.sheetnames}"
        )

    raw_rows = _read_sheet_rows(workbook, sheet_name)
    print(f"  Total rows in sheet (incl. header/blank): {len(raw_rows):,}")

    headers, data_rows = _extract_header_and_data(raw_rows)
    print(f"  Original column headers: {headers}")
    print(f"  Data rows after header detection: {len(data_rows):,}")

    if not data_rows:
        raise RuntimeError(f"No data rows found in {source_s3_uri} sheet '{sheet_name}'.")

    df = _build_spark_df(spark, headers, data_rows)

    if df.columns:
        agg_exprs = [
            F.sum(
                F.when(
                    F.col(c).isNotNull() & (F.length(F.trim(F.col(c))) > 0), 1
                ).otherwise(0)
            ).alias(c)
            for c in df.columns
        ]
        non_empty_counts = df.agg(*agg_exprs).collect()[0].asDict()
        null_or_empty_cols = [c for c, n in non_empty_counts.items() if (n or 0) == 0]
        if null_or_empty_cols:
            print(f"  Dropping {len(null_or_empty_cols)} fully-empty column(s): {null_or_empty_cols}")
            df = df.drop(*null_or_empty_cols)

    col_map = _dedupe_columns(df.columns)
    print("  Column rename map (original -> normalized):")
    for old, new in col_map.items():
        print(f"    {old!r:60s} -> {new}")
    for old, new in col_map.items():
        if old != new:
            df = df.withColumnRenamed(old, new)

    df = (
        df.withColumn("dt", F.lit(dt))
        .withColumn("_etl_loaded_at", F.lit(loaded_at))
        .withColumn("_source_s3_uri", F.lit(source_s3_uri))
        .withColumn("_source_folder", F.lit(folder_name))
        .withColumn("_source_sheet", F.lit(sheet_name))
    )

    table_location = output_s3_base.rstrip("/") + "/" + table_name + "/"

    if write_mode == "snapshot":
        print(f"  Snapshot mode: clearing previous contents at {table_location}")
        deleted = _delete_s3_prefix(table_location)
        print(f"  Deleted {deleted} previous object(s).")
        out_path = table_location
    else:
        out_path = table_location + f"dt={dt}/"

    print(f"  Writing curated data to: {out_path}")
    (
        df.coalesce(1)
        .write.mode("overwrite")
        .option("compression", "snappy")
        .parquet(out_path)
    )
    print("  Write complete.")

    final_schema_cols = [
        (f.name, _spark_type_to_hive(f.dataType))
        for f in df.schema.fields
        if not (write_mode == "daily_partition" and f.name == "dt")
    ]
    print(f"  Registering Glue table: {glue_database}.{table_name} (mode={write_mode})")
    _register_glue_table(
        database=glue_database,
        table=table_name,
        location=table_location,
        schema_columns=final_schema_cols,
        partitioned=(write_mode == "daily_partition"),
    )
    print("  Catalog registration complete.")


# =============================================================================
# Main
# =============================================================================

DEFAULT_FOLDER_TABLE_MAP = {
    "Ext_AbsoluteResolutionsInvestmentsLLC": "ext_absolute_resolutions_investments_llc",
    "Ext_DyckONeal": "ext_dyck_oneal",
    "Ext_Kirkland": "ext_kirkland",
    "Ext_LamarNationalBank": "ext_lamar_national_bank",
    "Ext_TMRR": "ext_tmrr",
}


args = getResolvedOptions(sys.argv, ["JOB_NAME"])

SOURCE_S3_PREFIX = _arg_or_default(
    "--SOURCE_S3_PREFIX",
    "s3://arrive-home-sftp-files/incoming/",
)
XLSX_FILENAME = _arg_or_default("--XLSX_FILENAME", "latest.xlsx")
OUTPUT_S3_BASE = _arg_or_default(
    "--OUTPUT_S3_BASE",
    "s3://arrivehome-bi-prod/curated/",
)
FOLDER_TABLE_MAP_JSON = _arg_or_default("--FOLDER_TABLE_MAP", "")
GLUE_DATABASE = _arg_or_default("--GLUE_DATABASE", "arrive_home")
DT = _arg_or_default("--DT", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
WRITE_MODE = _arg_or_default("--WRITE_MODE", "snapshot").strip().lower()
if WRITE_MODE not in ("snapshot", "daily_partition"):
    raise RuntimeError(
        f"Invalid --WRITE_MODE: {WRITE_MODE!r}. Expected 'snapshot' or 'daily_partition'."
    )

if FOLDER_TABLE_MAP_JSON.strip():
    try:
        FOLDER_TABLE_MAP = json.loads(FOLDER_TABLE_MAP_JSON)
        if not isinstance(FOLDER_TABLE_MAP, dict):
            raise ValueError("FOLDER_TABLE_MAP must be a JSON object")
    except Exception as e:
        raise RuntimeError(
            f"Invalid --FOLDER_TABLE_MAP JSON: {FOLDER_TABLE_MAP_JSON!r} ({e})"
        )
else:
    FOLDER_TABLE_MAP = DEFAULT_FOLDER_TABLE_MAP


sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args["JOB_NAME"], args)

spark.conf.set("spark.sql.parquet.datetimeRebaseModeInWrite", "CORRECTED")
spark.conf.set("spark.sql.parquet.int96RebaseModeInWrite", "CORRECTED")

print("=" * 80)
print("SFTP INCOMING XLSX → CURATED PARQUET (multi-folder)")
print("=" * 80)
print(f"Timestamp (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}")
print(f"DT: {DT}")
print(f"Source prefix: {SOURCE_S3_PREFIX}")
print(f"XLSX filename: {XLSX_FILENAME}")
print(f"Output base: {OUTPUT_S3_BASE}")
print(f"Glue DB: {GLUE_DATABASE}")
print(f"Write mode: {WRITE_MODE}")
print(f"Folder -> Table map: {json.dumps(FOLDER_TABLE_MAP, indent=2)}")

loaded_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
processed = 0

for folder_name, table_name in FOLDER_TABLE_MAP.items():
    print("\n" + "-" * 80)
    print(f"PROCESSING folder '{folder_name}' -> table '{GLUE_DATABASE}.{table_name}'")
    print("-" * 80)

    source_s3_uri = (
        SOURCE_S3_PREFIX.rstrip("/")
        + "/"
        + folder_name.strip("/")
        + "/"
        + XLSX_FILENAME
    )

    print(f"Downloading XLSX from {source_s3_uri} ...")
    xlsx_bytes = _load_xlsx_bytes_from_s3(source_s3_uri)
    print(f"  Downloaded {len(xlsx_bytes):,} bytes")

    workbook = _open_workbook(xlsx_bytes)
    sheet_names = workbook.sheetnames

    if not sheet_names:
        raise RuntimeError(f"No sheets found in {source_s3_uri}")

    # Each file has one tab — use the first sheet.
    sheet_name = sheet_names[0]
    print(f"  Using first sheet: '{sheet_name}'")

    _process_and_load(
        spark=spark,
        workbook=workbook,
        source_s3_uri=source_s3_uri,
        sheet_name=sheet_name,
        table_name=table_name,
        folder_name=folder_name,
        glue_database=GLUE_DATABASE,
        output_s3_base=OUTPUT_S3_BASE,
        write_mode=WRITE_MODE,
        dt=DT,
        loaded_at=loaded_at,
    )
    processed += 1

print(f"\nAll folders processed ({processed} table(s)).")
job.commit()
