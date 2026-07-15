"""
AWS Glue ETL Job: Servicing Dashboard XLSX sources → Curated Parquet (Athena)

Loads one or more .xlsx files from the servicing-dashboard-reporting bucket into
separate Athena tables in arrive_home (snapshot mode by default).

Default sources:
  - ServicingReporting.xlsx  -> dim_servicing_dashboard
  - USALoanTrialBal (.xls or .xlsx) -> dim_usa_loan_trial_bal

Each workbook has a single tab; the job uses the first sheet unless a sheet name
is specified in SOURCE_TABLE_MAP.

Required Glue job parameter (Job details → Advanced properties → Job parameters):
  Key:   --additional-python-modules
  Value: openpyxl==3.1.2,xlrd==2.0.1

If xlrd is still missing at runtime, the job will attempt a pip install to /tmp.

Parameters:
  --SOURCE_TABLE_MAP  (optional JSON array of {source_s3_uri, table, sheet?})
  --OUTPUT_S3_BASE    (default: s3://arrivehome-bi-prod/curated/)
  --GLUE_DATABASE     (default: arrive_home)
  --WRITE_MODE        (snapshot | daily_partition; default: snapshot)
  --DT                (default: today UTC)
"""

import io
import json
import os
import re
import subprocess
import sys
from datetime import date, datetime, timezone
from typing import Dict, List, Optional, Tuple

import boto3
from awsglue.context import GlueContext
from awsglue.job import Job
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql.types import StringType, StructField, StructType


# =============================================================================
# Helpers
# =============================================================================

def _arg_or_default(flag: str, default: str) -> str:
    if flag in sys.argv:
        return getResolvedOptions(sys.argv, [flag.lstrip("-")])[flag.lstrip("-")]
    return default


_word_re = re.compile(r"[A-Za-z0-9]+")
_camel_split_re = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")


def _to_snake_case(col: str) -> str:
    """Convert any header/text to Athena/Glue-friendly snake_case."""
    raw = (col or "").strip()
    if not raw:
        return "col"
    tokens = _word_re.findall(raw)
    words: List[str] = []
    for t in tokens:
        words.extend(p for p in _camel_split_re.split(t) if p)
    name = "_".join(w.lower() for w in words)
    if name and name[0].isdigit():
        name = f"col_{name}"
    return name or "col"


def _dedupe_columns(cols) -> Dict[str, str]:
    """Return mapping {original_col: new_unique_snake_case_col}."""
    seen: Dict[str, int] = {}
    mapping: Dict[str, str] = {}
    for col in cols:
        base = _to_snake_case(col)
        if base not in seen:
            seen[base] = 0
            mapping[col] = base
        else:
            seen[base] += 1
            mapping[col] = f"{base}_{seen[base]}"
    return mapping


def _parse_s3_uri(s3_uri: str) -> Tuple[str, str]:
    if not s3_uri.startswith("s3://"):
        raise ValueError(f"Expected s3:// URI, got: {s3_uri}")
    without = s3_uri[len("s3://") :]
    bucket, _, key = without.partition("/")
    if not bucket:
        raise ValueError(f"Expected s3://bucket/... URI, got: {s3_uri}")
    return bucket, key.lstrip("/")


def _is_numeric(v: str) -> bool:
    if not v:
        return False
    s = v.strip().replace(",", "").replace("$", "").replace("%", "")
    if s.startswith("-"):
        s = s[1:]
    s = s.replace(".", "", 1)
    return s.isdigit()


# =============================================================================
# Excel reader (openpyxl for .xlsx, xlrd for legacy .xls)
# =============================================================================

_GLUE_EXTRA_MODULES_DIR = "/tmp/glue_python_modules"


def _ensure_extra_modules_on_path() -> None:
    if _GLUE_EXTRA_MODULES_DIR not in sys.path:
        sys.path.insert(0, _GLUE_EXTRA_MODULES_DIR)


def _pip_install(package: str) -> None:
    os.makedirs(_GLUE_EXTRA_MODULES_DIR, exist_ok=True)
    print(f"  Installing {package} into {_GLUE_EXTRA_MODULES_DIR} ...")
    subprocess.check_call(
        [
            sys.executable,
            "-m",
            "pip",
            "install",
            package,
            "--target",
            _GLUE_EXTRA_MODULES_DIR,
            "--no-cache-dir",
        ]
    )
    _ensure_extra_modules_on_path()


def _import_xlrd():
    try:
        import xlrd

        return xlrd
    except ImportError:
        print("xlrd not found; attempting pip install ...")
        _pip_install("xlrd==2.0.1")
        import xlrd

        return xlrd


def _import_openpyxl():
    try:
        from openpyxl import load_workbook

        return load_workbook
    except ImportError:
        print("openpyxl not found; attempting pip install ...")
        _pip_install("openpyxl==3.1.2")
        from openpyxl import load_workbook

        return load_workbook


def _verify_excel_dependencies() -> None:
    print("Verifying Excel dependencies ...")
    xlrd = _import_xlrd()
    load_workbook = _import_openpyxl()
    print(f"  xlrd: {getattr(xlrd, '__version__', 'unknown')}")
    print(f"  openpyxl: OK (load_workbook={load_workbook})")


_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _detect_excel_format(file_bytes: bytes) -> str:
    """Detect real Excel format from file bytes (extension may lie after S3 copy)."""
    if file_bytes.startswith(_XLSX_MAGIC):
        return "xlsx"
    if file_bytes.startswith(_XLS_MAGIC):
        return "xls"
    raise RuntimeError(
        "Unrecognized Excel file format. Expected .xlsx (zip) or legacy .xls (OLE). "
        f"First bytes: {file_bytes[:8]!r}"
    )


def _load_xlsx_bytes_from_s3(s3_uri: str) -> bytes:
    bucket, key = _parse_s3_uri(s3_uri)
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    return obj["Body"].read()


_EXCEL_SUFFIXES = (".xlsx", ".xls")


def _find_latest_excel_by_stem(s3_prefix_uri: str, file_stem: str) -> str:
    """
    Find newest .xls or .xlsx under a prefix whose basename starts with file_stem.
    Example: USALoanTrialBal/latest/ -> USALoanTrialBal.xls or USALoanTrialBal.xlsx
    """
    bucket, prefix = _parse_s3_uri(s3_prefix_uri)
    if not prefix.endswith("/"):
        prefix = prefix + "/"

    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    matches = []
    stem_lower = file_stem.lower()

    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            basename = key.rsplit("/", 1)[-1]
            name_lower = basename.lower()
            if not name_lower.endswith(_EXCEL_SUFFIXES):
                continue
            if name_lower.startswith(stem_lower):
                matches.append(obj)

    if not matches:
        raise RuntimeError(
            f"No Excel file matching stem '{file_stem}' under s3://{bucket}/{prefix}"
        )

    latest = max(matches, key=lambda x: x["LastModified"])
    resolved = f"s3://{bucket}/{latest['Key']}"
    print(f"  Resolved {s3_prefix_uri} + stem '{file_stem}' -> {resolved}")
    return resolved


def _resolve_source_s3_uri(entry: Dict) -> str:
    if entry.get("source_s3_uri"):
        return entry["source_s3_uri"]
    prefix = entry.get("source_s3_prefix")
    stem = entry.get("source_file_stem")
    if prefix and stem:
        return _find_latest_excel_by_stem(prefix, stem)
    raise ValueError(
        "Each SOURCE_TABLE_MAP entry needs source_s3_uri or "
        "(source_s3_prefix + source_file_stem)"
    )


def _delete_s3_prefix(s3_uri: str) -> int:
    """
    Recursively delete every object under an s3:// prefix. Returns count deleted.

    Used in snapshot mode to wipe the previous table contents (including old
    dt=... folders left over from earlier partitioned runs) before writing.
    """
    bucket, prefix = _parse_s3_uri(s3_uri)
    if not prefix.endswith("/"):
        prefix = prefix + "/"
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    deleted = 0
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        objs = [{"Key": o["Key"]} for o in page.get("Contents", [])]
        if not objs:
            continue
        # delete_objects accepts up to 1000 keys per call
        for i in range(0, len(objs), 1000):
            s3.delete_objects(Bucket=bucket, Delete={"Objects": objs[i : i + 1000]})
        deleted += len(objs)
    return deleted


def _format_excel_date(dt: datetime) -> str:
    """Match Excel display for US loan trial balance reports (MM/DD/YYYY)."""
    return dt.strftime("%m/%d/%Y")


def _cell_to_string(value) -> str:
    """Convert a cell value to string, preserving Excel dates as MM/DD/YYYY."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return _format_excel_date(value)
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return str(value).strip()


def _xls_cell_value(sheet, rowx: int, colx: int, book) -> object:
    """Read one .xls cell; convert XL_CELL_DATE to MM/DD/YYYY string."""
    xlrd = _import_xlrd()
    ctype = sheet.cell_type(rowx, colx)
    value = sheet.cell_value(rowx, colx)

    if ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate.xldate_as_datetime(value, book.datemode)
            return _format_excel_date(dt)
        except Exception:
            try:
                y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(value, book.datemode)
                if (hh, mm, ss) == (0, 0, 0):
                    return f"{m:02d}/{d:02d}/{y}"
                return _format_excel_date(datetime(y, m, d, hh, mm, int(ss)))
            except Exception:
                return value

    if ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return str(int(value))
    if ctype == xlrd.XL_CELL_ERROR:
        return ""
    return value


def _normalize_xlsx_cell_value(value) -> object:
    if isinstance(value, datetime):
        return _format_excel_date(value)
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return value


def _is_likely_date_header(header: str) -> bool:
    h = (header or "").strip().lower()
    if h in ("ptd", "setup date", "paid to date"):
        return True
    return "date" in h


def _maybe_excel_serial_date(value, datemode: int = 0) -> object:
    """Convert Excel serial numbers in date columns (e.g. PTD 46174 -> 06/01/2026)."""
    if value is None or value == "":
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if 1 <= serial <= 295846:
            try:
                xlrd = _import_xlrd()
                dt = xlrd.xldate.xldate_as_datetime(serial, datemode)
                return _format_excel_date(dt)
            except Exception:
                pass
    return value


def _open_workbook(xlsx_bytes: bytes):
    load_workbook = _import_openpyxl()
    return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)


def _resolve_sheet_name(sheet_names: List[str], sheet_hint: Optional[str]) -> str:
    if sheet_hint and sheet_hint in sheet_names:
        return sheet_hint
    if sheet_hint and sheet_hint not in sheet_names:
        raise RuntimeError(
            f"Sheet '{sheet_hint}' not found. Available: {sheet_names}"
        )
    if not sheet_names:
        raise RuntimeError("Workbook has no sheets.")
    return sheet_names[0]


def _read_xls_sheet_rows(
    file_bytes: bytes, sheet_hint: Optional[str]
) -> Tuple[str, List[str], List[List[object]]]:
    xlrd = _import_xlrd()

    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet_names = book.sheet_names()
    sheet_name = _resolve_sheet_name(sheet_names, sheet_hint)
    sheet = book.sheet_by_name(sheet_name)
    rows = []
    for rowx in range(sheet.nrows):
        rows.append(
            [_xls_cell_value(sheet, rowx, colx, book) for colx in range(sheet.ncols)]
        )
    return sheet_name, sheet_names, rows


def _load_excel_sheet_rows(
    file_bytes: bytes, sheet_hint: Optional[str]
) -> Tuple[str, List[str], List[List[object]]]:
    """Load rows from .xlsx or legacy .xls (even if S3 key ends in .xlsx)."""
    fmt = _detect_excel_format(file_bytes)
    print(f"  Detected file format: {fmt}")

    if fmt == "xlsx":
        workbook = _open_workbook(file_bytes)
        sheet_names = workbook.sheetnames
        sheet_name = _resolve_sheet_name(sheet_names, sheet_hint)
        rows = _read_sheet_rows(workbook, sheet_name)
        return sheet_name, sheet_names, rows

    return _read_xls_sheet_rows(file_bytes, sheet_hint)


def _read_sheet_rows(workbook, sheet_name: str) -> List[List[object]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available: {workbook.sheetnames}"
        )
    sheet = workbook[sheet_name]
    rows: List[List[object]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_normalize_xlsx_cell_value(c) for c in row])
    return rows


def _detect_header_row(all_rows: List[List[object]]) -> int:
    """First row that's mostly populated AND mostly textual -> the real header."""
    if not all_rows:
        return 0
    max_cols = max(len(r) for r in all_rows)
    for i, row in enumerate(all_rows[:30]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for c in cells if c)
        if non_empty < max(3, max_cols // 2):
            continue
        text_cells = sum(1 for c in cells if c and not _is_numeric(c))
        if text_cells >= non_empty * 0.7:
            return i
    return 0


def _extract_header_and_data(
    all_rows: List[List[object]],
) -> Tuple[List[str], List[Tuple[str, ...]]]:
    """Return (original_headers, data_rows) with strings, padded to header width.

    Excel sheets often have formatting (borders, fills) extending thousands of
    columns past the real data. openpyxl will then report each row as having
    e.g. 16,000 columns. To stay sane we:
      1) Find the LAST header cell that has actual text.
      2) Truncate the header (and every data row) to that width.
    """
    if not all_rows:
        return [], []

    header_idx = _detect_header_row(all_rows)
    raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    # Find the rightmost non-empty header cell to use as the real column count.
    last_real_col = -1
    for i, h in enumerate(raw_headers):
        if h:
            last_real_col = i
    if last_real_col < 0:
        # Header row had no text at all (very unlikely after detection); keep as-is.
        last_real_col = len(raw_headers) - 1

    width = last_real_col + 1
    headers = raw_headers[:width]
    data = all_rows[header_idx + 1 :]

    date_col_indices = {i for i, h in enumerate(headers) if _is_likely_date_header(h)}
    if date_col_indices:
        print(f"  Date columns detected (by header): {[headers[i] for i in sorted(date_col_indices)]}")

    print(f"  Detected header row at xlsx row index: {header_idx} (0-based)")
    print(f"  Header reported width: {len(raw_headers)}; trimmed to: {width}")

    norm_rows: List[Tuple[str, ...]] = []
    for r in data:
        cells = []
        for i, c in enumerate(r[:width]):
            if i in date_col_indices:
                c = _maybe_excel_serial_date(c)
            cells.append(_cell_to_string(c))
        if len(cells) < width:
            cells = cells + [""] * (width - len(cells))
        if any(c.strip() for c in cells):
            norm_rows.append(tuple(cells))

    return headers, norm_rows


def _build_spark_df(spark, headers: List[str], data_rows: List[Tuple[str, ...]]) -> DataFrame:
    """Build a Spark DataFrame using the headers as columns; ensure unique names."""
    safe_headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(headers):
        base = h if h else f"col_{i}"
        if base not in seen:
            seen[base] = 0
            safe_headers.append(base)
        else:
            seen[base] += 1
            safe_headers.append(f"{base}_{seen[base]}")

    schema = StructType([StructField(h, StringType(), True) for h in safe_headers])
    return spark.createDataFrame(data_rows, schema=schema)


# =============================================================================
# Glue Catalog table registration (no crawler needed)
# =============================================================================

def _spark_type_to_hive(dtype) -> str:
    return dtype.simpleString()


def _register_glue_table(
    database: str,
    table: str,
    location: str,
    schema_columns: List[Tuple[str, str]],
    partitioned: bool = False,
) -> None:
    glue = boto3.client("glue")

    storage_descriptor = {
        "Columns": [{"Name": name, "Type": dtype} for name, dtype in schema_columns],
        "Location": location.rstrip("/") + "/",
        "InputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetInputFormat",
        "OutputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetOutputFormat",
        "Compressed": True,
        "SerdeInfo": {
            "SerializationLibrary": "org.apache.hadoop.hive.ql.io.parquet.serde.ParquetHiveSerDe",
            "Parameters": {"serialization.format": "1"},
        },
        "StoredAsSubDirectories": False,
    }

    parameters = {
        "classification": "parquet",
        "EXTERNAL": "TRUE",
    }
    partition_keys: List[Dict[str, str]] = []
    if partitioned:
        parameters.update(
            {
                "projection.enabled": "true",
                "projection.dt.type": "date",
                "projection.dt.format": "yyyy-MM-dd",
                "projection.dt.range": "2026-01-01,NOW",
                "projection.dt.interval": "1",
                "projection.dt.interval.unit": "DAYS",
                "storage.location.template": location.rstrip("/") + "/dt=${dt}",
            }
        )
        partition_keys = [{"Name": "dt", "Type": "string"}]

    table_input = {
        "Name": table,
        "TableType": "EXTERNAL_TABLE",
        "Parameters": parameters,
        "StorageDescriptor": storage_descriptor,
        "PartitionKeys": partition_keys,
    }

    try:
        glue.get_database(Name=database)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue database: {database}")
        glue.create_database(DatabaseInput={"Name": database})

    try:
        glue.get_table(DatabaseName=database, Name=table)
        print(f"  Updating Glue table: {database}.{table}")
        glue.update_table(DatabaseName=database, TableInput=table_input)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue table: {database}.{table}")
        glue.create_table(DatabaseName=database, TableInput=table_input)


# =============================================================================
# Main job
# =============================================================================

DEFAULT_SOURCE_TABLE_MAP = [
    {
        "source_s3_uri": (
            "s3://servicing-dashboard-reporting/ServicingDashboardReporting/latest/"
            "ServicingReporting.xlsx"
        ),
        "table": "dim_servicing_dashboard",
        "sheet": "Sheet1",
    },
    {
        "source_s3_prefix": "s3://servicing-dashboard-reporting/USALoanTrialBal/latest/",
        "source_file_stem": "USALoanTrialBal",
        "table": "dim_usa_loan_trial_bal",
    },
]


def _process_source(
    spark,
    source_s3_uri: str,
    table_name: str,
    sheet_hint: Optional[str],
    glue_database: str,
    output_s3_base: str,
    write_mode: str,
    dt: str,
    loaded_at: str,
) -> None:
    print(f"\nDownloading Excel from {source_s3_uri} ...")
    file_bytes = _load_xlsx_bytes_from_s3(source_s3_uri)
    print(f"  Downloaded {len(file_bytes):,} bytes")

    sheet_name, sheet_names, raw_rows = _load_excel_sheet_rows(file_bytes, sheet_hint)
    print(f"  Sheet names in workbook: {sheet_names}")
    if not sheet_hint:
        print(f"  Using first sheet: '{sheet_name}'")

    print("\n" + "-" * 80)
    print(f"PROCESSING '{source_s3_uri}' sheet '{sheet_name}' -> '{glue_database}.{table_name}'")
    print("-" * 80)

    print(f"  Total rows in sheet (incl. header/blank): {len(raw_rows):,}")

    headers, data_rows = _extract_header_and_data(raw_rows)
    print(f"  Original column headers: {headers}")
    print(f"  Data rows after header detection: {len(data_rows):,}")

    if not data_rows:
        raise RuntimeError(f"No data rows in {source_s3_uri} sheet '{sheet_name}'.")

    df = _build_spark_df(spark, headers, data_rows)

    if df.columns:
        agg_exprs = [
            F.sum(
                F.when(
                    F.col(c).isNotNull() & (F.length(F.trim(F.col(c))) > 0), 1
                ).otherwise(0)
            ).alias(c)
            for c in df.columns
        ]
        non_empty_counts = df.agg(*agg_exprs).collect()[0].asDict()
        null_or_empty_cols = [c for c, n in non_empty_counts.items() if (n or 0) == 0]
        if null_or_empty_cols:
            print(f"  Dropping {len(null_or_empty_cols)} fully-empty column(s): {null_or_empty_cols}")
            df = df.drop(*null_or_empty_cols)

    col_map = _dedupe_columns(df.columns)
    print("  Column rename map (original -> normalized):")
    for old, new in col_map.items():
        print(f"    {old!r:60s} -> {new}")
    for old, new in col_map.items():
        if old != new:
            df = df.withColumnRenamed(old, new)

    df = (
        df.withColumn("dt", F.lit(dt))
        .withColumn("_etl_loaded_at", F.lit(loaded_at))
        .withColumn("_source_s3_uri", F.lit(source_s3_uri))
        .withColumn("_source_sheet", F.lit(sheet_name))
    )

    table_location = output_s3_base.rstrip("/") + "/" + table_name + "/"

    if write_mode == "snapshot":
        print(f"  Snapshot mode: clearing previous contents at {table_location}")
        deleted = _delete_s3_prefix(table_location)
        print(f"  Deleted {deleted} previous object(s).")
        out_path = table_location
    else:
        out_path = table_location + f"dt={dt}/"

    print(f"  Writing curated data to: {out_path}")
    (
        df.coalesce(1)
        .write.mode("overwrite")
        .option("compression", "snappy")
        .parquet(out_path)
    )
    print("  Write complete.")

    final_schema_cols = [
        (f.name, _spark_type_to_hive(f.dataType))
        for f in df.schema.fields
        if not (write_mode == "daily_partition" and f.name == "dt")
    ]
    print(f"  Registering Glue table: {glue_database}.{table_name} (mode={write_mode})")
    _register_glue_table(
        database=glue_database,
        table=table_name,
        location=table_location,
        schema_columns=final_schema_cols,
        partitioned=(write_mode == "daily_partition"),
    )
    print("  Catalog registration complete.")


def _parse_source_table_map(source_table_map_json: str) -> List[Dict]:
    if source_table_map_json.strip():
        try:
            parsed = json.loads(source_table_map_json)
        except Exception as e:
            raise RuntimeError(
                f"Invalid --SOURCE_TABLE_MAP JSON: {source_table_map_json!r} ({e})"
            )
        if not isinstance(parsed, list):
            raise ValueError("SOURCE_TABLE_MAP must be a JSON array")
        for i, entry in enumerate(parsed):
            if not isinstance(entry, dict):
                raise ValueError(f"SOURCE_TABLE_MAP[{i}] must be an object")
            if "table" not in entry:
                raise ValueError(f"SOURCE_TABLE_MAP[{i}] must include table")
            has_uri = bool(entry.get("source_s3_uri"))
            has_prefix_stem = bool(entry.get("source_s3_prefix")) and bool(
                entry.get("source_file_stem")
            )
            if not has_uri and not has_prefix_stem:
                raise ValueError(
                    f"SOURCE_TABLE_MAP[{i}] needs source_s3_uri or "
                    "(source_s3_prefix + source_file_stem)"
                )
        return parsed
    return DEFAULT_SOURCE_TABLE_MAP


args = getResolvedOptions(sys.argv, ["JOB_NAME"])

OUTPUT_S3_BASE = _arg_or_default(
    "--OUTPUT_S3_BASE",
    "s3://arrivehome-bi-prod/curated/",
)
SOURCE_TABLE_MAP_JSON = _arg_or_default("--SOURCE_TABLE_MAP", "")
GLUE_DATABASE = _arg_or_default("--GLUE_DATABASE", "arrive_home")
DT = _arg_or_default("--DT", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
# WRITE_MODE controls how repeated runs behave:
#   snapshot         (default) -> each run REPLACES the table data (no duplicates).
#                                 dt is kept as a regular column for traceability.
#   daily_partition           -> each run writes to dt=YYYY-MM-DD/ (history kept).
#                                 Athena queries must filter by dt to avoid dupes.
WRITE_MODE = _arg_or_default("--WRITE_MODE", "snapshot").strip().lower()
if WRITE_MODE not in ("snapshot", "daily_partition"):
    raise RuntimeError(
        f"Invalid --WRITE_MODE: {WRITE_MODE!r}. Expected 'snapshot' or 'daily_partition'."
    )

SOURCE_TABLE_MAP = _parse_source_table_map(SOURCE_TABLE_MAP_JSON)


sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args["JOB_NAME"], args)

_verify_excel_dependencies()

spark.conf.set("spark.sql.parquet.datetimeRebaseModeInWrite", "CORRECTED")
spark.conf.set("spark.sql.parquet.int96RebaseModeInWrite", "CORRECTED")

print("=" * 80)
print("SERVICING DASHBOARD XLSX → CURATED PARQUET (multi-source)")
print("=" * 80)
print(f"Timestamp (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}")
print(f"DT: {DT}")
print(f"Output base: {OUTPUT_S3_BASE}")
print(f"Glue DB: {GLUE_DATABASE}")
print(f"Write mode: {WRITE_MODE}")
print(f"Source -> Table map: {json.dumps(SOURCE_TABLE_MAP, indent=2)}")

loaded_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
processed = 0

for entry in SOURCE_TABLE_MAP:
    source_s3_uri = _resolve_source_s3_uri(entry)
    _process_source(
        spark=spark,
        source_s3_uri=source_s3_uri,
        table_name=entry["table"],
        sheet_hint=entry.get("sheet"),
        glue_database=GLUE_DATABASE,
        output_s3_base=OUTPUT_S3_BASE,
        write_mode=WRITE_MODE,
        dt=DT,
        loaded_at=loaded_at,
    )
    processed += 1

print(f"\nAll sources processed ({processed} table(s)).")
job.commit()
