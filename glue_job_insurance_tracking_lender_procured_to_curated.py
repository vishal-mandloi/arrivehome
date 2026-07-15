"""
AWS Glue ETL Job: Insurance Tracking XLSX (multiple sheets) → Curated Parquet (Athena)

Source: S3 .xlsx file written daily by Amazon AppFlow (SharePoint → S3)
Input example: s3://arrive-home-insurance-tracking/InsuranceTrackingSharepoint/latest/InsuranceTracking.xlsx

Output: S3 curated zone (Parquet, Snappy), partitioned by dt=YYYY-MM-DD
Each Excel sheet is written to its own subfolder + its own Glue Catalog table.

Default sheet → table mapping (override via --SHEET_TABLE_MAP):
  - "Lender procured"                  -> insurance_tracking_lender_procured
  - "Unification of AH procured prem"  -> insurance_tracking_ah_procured

Derived curated tables (built after both sheets load):
  - fact_procurement   (union + normalization; SourceType Internal/External)
  - dim_loan_master    (one row per Loan_Key_v3, Power BI parity)
  - dim_month          (distinct Request/Close month starts from fact_procurement)
  - dim_date           (calendar 2024-01-01 .. 2027-12-31)

Required Glue job parameter (Glue Studio: "Job parameters"):
  --additional-python-modules openpyxl

Other parameters:
  --SOURCE_S3_URI    (default: the SharePoint xlsx in arrive-home-insurance-tracking)
  --OUTPUT_S3_BASE   (default: s3://arrivehome-bi-prod/curated/)
  --SHEET_TABLE_MAP  (optional JSON: {"Sheet name": "athena_table_name", ...})
  --GLUE_DATABASE    (default: insurance_tracking_sharepoint)
  --DT               (default: today UTC, "YYYY-MM-DD")
"""

import io
import json
import re
import sys
from datetime import date, datetime, timezone
from typing import Dict, List, Optional, Tuple

import boto3
from awsglue.context import GlueContext
from awsglue.job import Job
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql.types import StringType, StructField, StructType


# =============================================================================
# Helpers
# =============================================================================

def _arg_or_default(flag: str, default: str) -> str:
    if flag in sys.argv:
        return getResolvedOptions(sys.argv, [flag.lstrip("-")])[flag.lstrip("-")]
    return default


_word_re = re.compile(r"[A-Za-z0-9]+")
_camel_split_re = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")


def _to_snake_case(col: str) -> str:
    """Convert any header/text to Athena/Glue-friendly snake_case."""
    raw = (col or "").strip()
    if not raw:
        return "col"
    tokens = _word_re.findall(raw)
    words: List[str] = []
    for t in tokens:
        words.extend(p for p in _camel_split_re.split(t) if p)
    name = "_".join(w.lower() for w in words)
    if name and name[0].isdigit():
        name = f"col_{name}"
    return name or "col"


def _dedupe_columns(cols) -> Dict[str, str]:
    """Return mapping {original_col: new_unique_snake_case_col}."""
    seen: Dict[str, int] = {}
    mapping: Dict[str, str] = {}
    for col in cols:
        base = _to_snake_case(col)
        if base not in seen:
            seen[base] = 0
            mapping[col] = base
        else:
            seen[base] += 1
            mapping[col] = f"{base}_{seen[base]}"
    return mapping


def _parse_s3_uri(s3_uri: str) -> Tuple[str, str]:
    if not s3_uri.startswith("s3://"):
        raise ValueError(f"Expected s3:// URI, got: {s3_uri}")
    without = s3_uri[len("s3://") :]
    bucket, _, key = without.partition("/")
    if not bucket:
        raise ValueError(f"Expected s3://bucket/... URI, got: {s3_uri}")
    return bucket, key.lstrip("/")


def _is_numeric(v: str) -> bool:
    if not v:
        return False
    s = v.strip().replace(",", "").replace("$", "").replace("%", "")
    if s.startswith("-"):
        s = s[1:]
    s = s.replace(".", "", 1)
    return s.isdigit()


# =============================================================================
# XLSX reader (openpyxl)
# =============================================================================

def _load_xlsx_bytes_from_s3(s3_uri: str) -> bytes:
    bucket, key = _parse_s3_uri(s3_uri)
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    return obj["Body"].read()


def _delete_s3_prefix(s3_uri: str) -> int:
    """
    Recursively delete every object under an s3:// prefix. Returns count deleted.

    Used in snapshot mode to wipe the previous table contents (including old
    dt=... folders left over from earlier partitioned runs) before writing.
    """
    bucket, prefix = _parse_s3_uri(s3_uri)
    if not prefix.endswith("/"):
        prefix = prefix + "/"
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    deleted = 0
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        objs = [{"Key": o["Key"]} for o in page.get("Contents", [])]
        if not objs:
            continue
        # delete_objects accepts up to 1000 keys per call
        for i in range(0, len(objs), 1000):
            s3.delete_objects(Bucket=bucket, Delete={"Objects": objs[i : i + 1000]})
        deleted += len(objs)
    return deleted


def _open_workbook(xlsx_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required. Add `--additional-python-modules openpyxl` "
            "to the Glue job parameters."
        ) from e
    return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)


def _read_sheet_rows(workbook, sheet_name: str) -> List[List[object]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available: {workbook.sheetnames}"
        )
    sheet = workbook[sheet_name]
    rows: List[List[object]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _detect_header_row(all_rows: List[List[object]]) -> int:
    """First row that's mostly populated AND mostly textual -> the real header."""
    if not all_rows:
        return 0
    max_cols = max(len(r) for r in all_rows)
    for i, row in enumerate(all_rows[:30]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for c in cells if c)
        if non_empty < max(3, max_cols // 2):
            continue
        text_cells = sum(1 for c in cells if c and not _is_numeric(c))
        if text_cells >= non_empty * 0.7:
            return i
    return 0


def _extract_header_and_data(
    all_rows: List[List[object]],
) -> Tuple[List[str], List[Tuple[str, ...]]]:
    """Return (original_headers, data_rows) with strings, padded to header width.

    Excel sheets often have formatting (borders, fills) extending thousands of
    columns past the real data. openpyxl will then report each row as having
    e.g. 16,000 columns. To stay sane we:
      1) Find the LAST header cell that has actual text.
      2) Truncate the header (and every data row) to that width.
    """
    if not all_rows:
        return [], []

    header_idx = _detect_header_row(all_rows)
    raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    # Find the rightmost non-empty header cell to use as the real column count.
    last_real_col = -1
    for i, h in enumerate(raw_headers):
        if h:
            last_real_col = i
    if last_real_col < 0:
        # Header row had no text at all (very unlikely after detection); keep as-is.
        last_real_col = len(raw_headers) - 1

    width = last_real_col + 1
    headers = raw_headers[:width]
    data = all_rows[header_idx + 1 :]

    print(f"  Detected header row at xlsx row index: {header_idx} (0-based)")
    print(f"  Header reported width: {len(raw_headers)}; trimmed to: {width}")

    norm_rows: List[Tuple[str, ...]] = []
    for r in data:
        cells = [("" if c is None else str(c)) for c in r[:width]]
        if len(cells) < width:
            cells = cells + [""] * (width - len(cells))
        if any(c.strip() for c in cells):
            norm_rows.append(tuple(cells))

    return headers, norm_rows


def _build_spark_df(spark, headers: List[str], data_rows: List[Tuple[str, ...]]) -> DataFrame:
    """Build a Spark DataFrame using the headers as columns; ensure unique names."""
    safe_headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(headers):
        base = h if h else f"col_{i}"
        if base not in seen:
            seen[base] = 0
            safe_headers.append(base)
        else:
            seen[base] += 1
            safe_headers.append(f"{base}_{seen[base]}")

    schema = StructType([StructField(h, StringType(), True) for h in safe_headers])
    return spark.createDataFrame(data_rows, schema=schema)


# =============================================================================
# Glue Catalog table registration (no crawler needed)
# =============================================================================

def _spark_type_to_hive(dtype) -> str:
    return dtype.simpleString()


def _register_glue_table(
    database: str,
    table: str,
    location: str,
    schema_columns: List[Tuple[str, str]],
    partitioned: bool = False,
) -> None:
    glue = boto3.client("glue")

    storage_descriptor = {
        "Columns": [{"Name": name, "Type": dtype} for name, dtype in schema_columns],
        "Location": location.rstrip("/") + "/",
        "InputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetInputFormat",
        "OutputFormat": "org.apache.hadoop.hive.ql.io.parquet.MapredParquetOutputFormat",
        "Compressed": True,
        "SerdeInfo": {
            "SerializationLibrary": "org.apache.hadoop.hive.ql.io.parquet.serde.ParquetHiveSerDe",
            "Parameters": {"serialization.format": "1"},
        },
        "StoredAsSubDirectories": False,
    }

    parameters = {
        "classification": "parquet",
        "EXTERNAL": "TRUE",
    }
    partition_keys: List[Dict[str, str]] = []
    if partitioned:
        parameters.update(
            {
                "projection.enabled": "true",
                "projection.dt.type": "date",
                "projection.dt.format": "yyyy-MM-dd",
                "projection.dt.range": "2026-01-01,NOW",
                "projection.dt.interval": "1",
                "projection.dt.interval.unit": "DAYS",
                "storage.location.template": location.rstrip("/") + "/dt=${dt}",
            }
        )
        partition_keys = [{"Name": "dt", "Type": "string"}]

    table_input = {
        "Name": table,
        "TableType": "EXTERNAL_TABLE",
        "Parameters": parameters,
        "StorageDescriptor": storage_descriptor,
        "PartitionKeys": partition_keys,
    }

    try:
        glue.get_database(Name=database)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue database: {database}")
        glue.create_database(DatabaseInput={"Name": database})

    try:
        glue.get_table(DatabaseName=database, Name=table)
        print(f"  Updating Glue table: {database}.{table}")
        glue.update_table(DatabaseName=database, TableInput=table_input)
    except glue.exceptions.EntityNotFoundException:
        print(f"  Creating Glue table: {database}.{table}")
        glue.create_table(DatabaseName=database, TableInput=table_input)


# =============================================================================
# fact_procurement + dimension builders (Power BI / QuickSight parity)
# =============================================================================

def _col_map(df: DataFrame) -> Dict[str, str]:
    return {c.lower(): c for c in df.columns}


def _pick_col(df: DataFrame, col_map: Dict[str, str], *candidates: str):
    for cand in candidates:
        key = cand.lower()
        if key in col_map:
            return F.col(col_map[key])
    return F.lit(None).cast("string")


def _norm_alnum_token(expr) -> F.Column:
    """Uppercase + keep 0-9/A-Z only; null when empty (Power Query Text.Select)."""
    kept = F.upper(
        F.regexp_replace(
            F.trim(F.coalesce(expr.cast("string"), F.lit(""))),
            r"[^A-Z0-9]",
            "",
        )
    )
    return F.when(F.length(kept) == 0, F.lit(None).cast("string")).otherwise(kept)


def _norm_lender_name(expr) -> F.Column:
    raw = F.upper(F.trim(F.coalesce(expr.cast("string"), F.lit(""))))
    s = raw
    for token in (
        "ITS SUCCESSORS AND/OR ASSIGNS",
        "AND/OR ASSIGNS",
        "ITS SUCCESSORS",
        "ISAOA",
        "ATIMA",
    ):
        s = F.regexp_replace(s, token, " ")
    kept = F.regexp_replace(s, r"[^A-Z0-9 ]", "")
    collapsed = F.regexp_replace(F.trim(kept), r" +", " ")
    return F.when(F.length(collapsed) == 0, F.lit(None).cast("string")).otherwise(
        collapsed
    )


def _norm_address_norm(expr) -> F.Column:
    raw = F.upper(F.trim(F.coalesce(expr.cast("string"), F.lit(""))))
    kept = F.regexp_replace(raw, r"[^A-Z0-9 ]", "")
    collapsed = F.regexp_replace(F.trim(kept), r" +", " ")
    return F.when(F.length(collapsed) == 0, F.lit(None).cast("string")).otherwise(
        collapsed
    )


def _zip_norm(expr) -> F.Column:
    digits = F.regexp_replace(
        F.coalesce(expr.cast("string"), F.lit("")), r"[^0-9]", ""
    )
    return F.when(F.length(digits) >= 5, F.substring(digits, 1, 5)).otherwise(
        F.lit(None).cast("string")
    )


def _parse_flexible_date(expr) -> F.Column:
    """Parse Excel/openpyxl date strings into Spark date; invalid -> null.

    openpyxl often stringifies cells as '2024-11-18 00:00:00'. Malformed values
    like '07/01/2026/2026' must become blank — never fail the job. Spark's to_date
    throws on bad input, and coalesce(to_date(...), ...) still evaluates every
    branch, so we use try_to_date when available or regex-gated to_date.
    """
    s = F.trim(F.coalesce(expr.cast("string"), F.lit("")))
    invalid = (
        s.isNull()
        | (F.length(s) == 0)
        | F.lower(s).rlike(
            r"not going|pending|cancel|n/a|na|tbd|unknown|none|^-$"
        )
    )
    # Drop " 00:00:00", "T00:00:00", fractional seconds, etc.
    s_date = F.regexp_replace(
        s,
        r"([Tt]|\s+)\d{1,2}:\d{2}(:\d{2})?(\.\d+)?$",
        "",
    )

    try_to_date = getattr(F, "try_to_date", None)
    if try_to_date is not None:
        # try_to_date never throws — bad strings become null.
        parsed = F.coalesce(
            try_to_date(s_date, "yyyy-MM-dd"),
            try_to_date(s_date, "M/d/yyyy"),
            try_to_date(s_date, "MM/dd/yyyy"),
            try_to_date(s_date, "M/d/yy"),
            try_to_date(s_date, "MM/dd/yy"),
        )
    else:
        # Regex-gated to_date: only parse when the whole string matches a format.
        parsed = (
            F.when(
                s_date.rlike(r"^\d{4}-\d{2}-\d{2}$"),
                F.to_date(s_date, "yyyy-MM-dd"),
            )
            .when(
                s_date.rlike(r"^\d{1,2}/\d{1,2}/\d{4}$"),
                F.to_date(s_date, "M/d/yyyy"),
            )
            .when(
                s_date.rlike(r"^\d{1,2}/\d{1,2}/\d{2}$"),
                F.to_date(s_date, "M/d/yy"),
            )
            .otherwise(F.lit(None).cast("date"))
        )

    return F.when(invalid, F.lit(None).cast("date")).otherwise(parsed)


def _parse_money(expr) -> F.Column:
    cleaned = F.regexp_replace(
        F.coalesce(expr.cast("string"), F.lit("")), r"[$,\s]", ""
    )
    return F.when(F.length(cleaned) == 0, F.lit(None).cast("double")).otherwise(
        cleaned.cast("double")
    )


def _standardize_procurement_source(
    df: DataFrame, source_type: str
) -> DataFrame:
    """Map either raw sheet into stg_*-like columns before union."""
    cm = _col_map(df)
    return df.select(
        F.lit(source_type).alias("source_type"),
        _pick_col(df, cm, "borrower_first_name", "first_name").alias("first_name"),
        _pick_col(df, cm, "borrower_last_name", "last_name").alias("last_name"),
        _pick_col(df, cm, "home_buyers_name", "homebuyers_name").alias(
            "home_buyers_name"
        ),
        _pick_col(
            df,
            cm,
            "lender_loan",
            "lender_loan_number",
            "lender_loan_no",
            "lender_loan_number_raw",
        ).alias("lender_loan_number"),
        _pick_col(
            df,
            cm,
            "arrive_home_loan_number",
            "ah_loan_number",
            "ah_loan",
        ).alias("arrive_home_loan_number"),
        _pick_col(
            df,
            cm,
            "certificate_policy",
            "certificate_policy_number",
            "policy_number",
            "certificate_policy_no",
        ).alias("certificate_policy_number"),
        _pick_col(df, cm, "lender", "lender_name").alias("lender_name"),
        _pick_col(
            df, cm, "borrower_address", "address", "property_address"
        ).alias("address"),
        _pick_col(df, cm, "borrower_city", "city", "property_city").alias("city"),
        _pick_col(df, cm, "borrower_state", "state", "property_state").alias("state"),
        _pick_col(df, cm, "borrower_zip", "zip", "property_zip").alias("zip"),
        _pick_col(df, cm, "county").alias("county"),
        _pick_col(
            df,
            cm,
            "date_requested_by_lender",
            "request_date",
            "date_requested",
        ).alias("request_date_raw"),
        _pick_col(
            df, cm, "closed_bind", "close_bind_date", "closed_bind_date"
        ).alias("closed_bind_raw"),
        _pick_col(df, cm, "effective_date").alias("effective_date_raw"),
        _pick_col(df, cm, "approval_status").alias("approval_status_raw"),
        _pick_col(df, cm, "policy_type", "policy_type_1").alias("policy_type_raw"),
        _pick_col(df, cm, "policy_type_secondary", "policy_type_2").alias(
            "policy_type_secondary_raw"
        ),
        _pick_col(df, cm, "premium").alias("premium_raw"),
        _pick_col(df, cm, "flood_required").alias("flood_required_raw"),
        _pick_col(df, cm, "processor").alias("processor"),
        _pick_col(df, cm, "insurance_rep").alias("insurance_rep"),
        _pick_col(df, cm, "payment_status", "status").alias("status_raw"),
        _pick_col(df, cm, "property_conditions", "property_condition").alias(
            "property_condition_raw"
        ),
        _pick_col(df, cm, "year", "year_built").alias("year_built_raw"),
        _pick_col(
            df, cm, "gross_living_area_sq_ft", "sqft", "square_feet"
        ).alias("sqft"),
        _pick_col(df, cm, "fema_flood_zone", "flood_zone").alias("flood_zone"),
        _pick_col(
            df, cm, "dwelling_coverage", "lender_procured_rce", "dwelling_amount"
        ).alias("dwelling_amount"),
        _pick_col(df, cm, "carrier_s", "carrier").alias("carrier"),
        _pick_col(df, cm, "broker").alias("broker"),
        _pick_col(df, cm, "provider").alias("provider"),
        _pick_col(
            df,
            cm,
            "sent_to_procurement_date",
            "date_sent_to_procurement",
        ).alias("sent_to_procurement_date_raw"),
        _pick_col(
            df, cm, "cert_received_date", "certificate_received_date"
        ).alias("cert_received_date_raw"),
        _pick_col(df, cm, "follow_up_date", "followup_date").alias(
            "follow_up_date_raw"
        ),
        _pick_col(df, cm, "determination_date").alias(
            "determination_date_raw"
        ),
    )


def _build_fact_procurement(
    lender_df: Optional[DataFrame], ah_df: Optional[DataFrame]
) -> Optional[DataFrame]:
    parts: List[DataFrame] = []
    if lender_df is not None:
        parts.append(_standardize_procurement_source(lender_df, "External"))
    if ah_df is not None:
        parts.append(
            _standardize_procurement_source(ah_df, "Internal Procured")
        )
    if not parts:
        return None

    combined = parts[0]
    for part in parts[1:]:
        combined = combined.unionByName(part, allowMissingColumns=True)

    borrower_name = F.trim(
        F.concat_ws(
            " ",
            F.nullif(F.trim(F.coalesce(F.col("first_name"), F.lit(""))), F.lit("")),
            F.nullif(F.trim(F.coalesce(F.col("last_name"), F.lit(""))), F.lit("")),
        )
    )
    home_buyers_upper = F.upper(
        F.trim(F.coalesce(F.col("home_buyers_name").cast("string"), F.lit("")))
    )
    borrower_upper = F.upper(F.trim(F.coalesce(borrower_name, F.lit(""))))
    loan_key = (
        F.when(F.length(home_buyers_upper) > 0, home_buyers_upper)
        .when(F.length(borrower_upper) > 0, borrower_upper)
        .otherwise(F.lit(None).cast("string"))
    )

    request_date = _parse_flexible_date(F.col("request_date_raw"))
    close_bind_date = _parse_flexible_date(F.col("closed_bind_raw"))
    effective_date = _parse_flexible_date(F.col("effective_date_raw"))
    sent_to_procurement_date = _parse_flexible_date(
        F.col("sent_to_procurement_date_raw")
    )
    cert_received_date = _parse_flexible_date(F.col("cert_received_date_raw"))
    follow_up_date = _parse_flexible_date(F.col("follow_up_date_raw"))
    determination_date = _parse_flexible_date(F.col("determination_date_raw"))

    closed_flag = F.when(close_bind_date.isNotNull(), F.lit(1)).otherwise(F.lit(0))

    src = F.upper(F.trim(F.coalesce(F.col("source_type"), F.lit(""))))
    s1 = F.upper(F.trim(F.coalesce(F.col("status_raw"), F.lit(""))))
    s2 = F.upper(F.trim(F.coalesce(F.col("approval_status_raw"), F.lit(""))))
    is_internal = src.contains("INTERNAL") | src.contains("UNIFICATION")
    status_internal = (
        F.when(s1 == "ACTIVE", F.lit("ACTIVE"))
        .when(s1 == "PENDING", F.lit("PENDING"))
        .when(s1 == "CANCELLED", F.lit("CANCELLED"))
        .when(s1 == "", F.lit("BLANK"))
        .otherwise(s1)
    )
    status_external = (
        F.when(s2 == "APPROVED", F.lit("APPROVED"))
        .when(s2 == "CLOSED WITHOUT APPROVAL", F.lit("CLOSED WITHOUT APPROVAL"))
        .when(s2 == "CANCELLED", F.lit("CANCELLED"))
        .when(s2 == "DENIED", F.lit("DENIED"))
        .when(s2 == "PENDING ITEMS", F.lit("PENDING"))
        .when(s2 == "", F.when(s1 == "", F.lit("BLANK")).otherwise(s1))
        .otherwise(s2)
    )
    status_normalized = F.when(is_internal, status_internal).otherwise(
        status_external
    )

    flood_raw = F.upper(F.trim(F.coalesce(F.col("flood_required_raw"), F.lit(""))))
    flood_required_flag = (
        F.when(flood_raw == "", F.lit(0))
        .when(
            (flood_raw == "NO")
            | (flood_raw == "N")
            | (flood_raw == "FALSE")
            | flood_raw.contains("NOT REQUIRED"),
            F.lit(0),
        )
        .when(
            (flood_raw == "YES")
            | (flood_raw == "Y")
            | (flood_raw == "TRUE")
            | flood_raw.contains("REQUIRED"),
            F.lit(1),
        )
        .otherwise(F.lit(0))
    )
    is_flood_policy = flood_required_flag
    is_hazard_policy = F.when(flood_required_flag == 0, F.lit(1)).otherwise(F.lit(0))

    property_raw = F.trim(
        F.coalesce(F.col("property_condition_raw").cast("string"), F.lit(""))
    )
    property_upper = F.upper(property_raw)
    property_condition = (
        F.when(property_upper.isin("C1", "C2", "C3", "C4", "C5"), property_upper)
        .when(F.length(property_raw) == 0, F.lit("MNF - BLANK"))
        .otherwise(F.concat(F.lit("MNF - "), property_raw))
    )

    year_built = F.coalesce(
        F.col("year_built_raw").cast("int"),
        F.regexp_extract(
            F.coalesce(F.col("year_built_raw").cast("string"), F.lit("")),
            r"(\d{4})",
            1,
        ).cast("int"),
    )
    age_bucket = (
        F.when(year_built.isNull(), F.lit(None).cast("string"))
        .when(year_built <= 1978, F.lit("1978 or older"))
        .otherwise(F.lit("Newer than 1978"))
    )

    request_month_start = F.when(
        request_date.isNotNull(), F.trunc(request_date, "month")
    ).otherwise(F.lit(None).cast("date"))
    close_month_start = F.when(
        close_bind_date.isNotNull(), F.trunc(close_bind_date, "month")
    ).otherwise(F.lit(None).cast("date"))
    days_to_close = F.when(
        request_date.isNotNull() & close_bind_date.isNotNull(),
        F.datediff(close_bind_date, request_date),
    ).otherwise(F.lit(None).cast("int"))

    loan_number_norm = _norm_alnum_token(F.col("lender_loan_number"))
    arrive_home_loan_number_norm = _norm_alnum_token(F.col("arrive_home_loan_number"))
    policy_number_norm = _norm_alnum_token(F.col("certificate_policy_number"))

    loan_key_v2 = (
        F.when(
            loan_number_norm.isNotNull(),
            F.concat(F.lit("LN:"), loan_number_norm),
        )
        .when(
            loan_key.isNotNull() & (F.length(F.trim(loan_key)) > 0),
            F.concat(F.lit("NAME:"), loan_key),
        )
        .otherwise(F.lit(None).cast("string"))
    )

    flood_required_code = (
        F.when(flood_raw == "", F.lit("UNK"))
        .when(
            (flood_raw == "NO")
            | (flood_raw == "N")
            | (flood_raw == "FALSE")
            | flood_raw.contains("NOT REQUIRED"),
            F.lit("NOT_REQ"),
        )
        .when(
            (flood_raw == "YES")
            | (flood_raw == "Y")
            | (flood_raw == "TRUE")
            | flood_raw.contains("REQUIRED"),
            F.lit("REQ"),
        )
        .otherwise(F.lit("UNK"))
    )
    flood_required_label = (
        F.when(flood_required_code == "REQ", F.lit("Flood Required"))
        .when(flood_required_code == "NOT_REQ", F.lit("No Flood Required"))
        .otherwise(F.lit("Unknown"))
    )

    status_norm_upper = F.upper(F.trim(F.coalesce(status_normalized, F.lit(""))))
    status_normalized_v2 = (
        F.when(status_norm_upper == "ACTIVE", F.lit("Active"))
        .when(status_norm_upper == "PENDING", F.lit("Pending"))
        .when(status_norm_upper == "APPROVED", F.lit("Approved"))
        .when(
            status_norm_upper.isin("CANCELLED", "CANCELED"),
            F.lit("Cancelled"),
        )
        .when(status_norm_upper == "DENIED", F.lit("Denied"))
        .when(
            status_norm_upper == "CLOSED WITHOUT APPROVAL",
            F.lit("Closed Without Approval"),
        )
        .when(
            status_norm_upper.contains("PAID") & status_norm_upper.contains("FULL"),
            F.lit("Paid In Full"),
        )
        .when(
            status_norm_upper.isin("BLANK", ""),
            F.lit("Blank"),
        )
        .otherwise(F.lit("Unmapped"))
    )
    status_mapped = F.when(status_normalized_v2 == "Unmapped", F.lit(0)).otherwise(
        F.lit(1)
    )

    lender_name_norm = _norm_lender_name(F.col("lender_name"))
    address_norm = _norm_address_norm(F.col("address"))
    zip_norm = _zip_norm(F.col("zip"))
    address_a = F.coalesce(address_norm, F.lit(""))
    address_z = F.coalesce(zip_norm, F.lit(""))
    address_key = F.when(
        (F.length(address_a) == 0) & (F.length(address_z) == 0),
        F.lit(None).cast("string"),
    ).otherwise(F.concat_ws("|", address_a, address_z))

    lender_for_key = F.coalesce(lender_name_norm, F.lit(""))
    addr_for_key = F.coalesce(address_key, F.lit(""))
    loan_key_v3 = (
        F.when(
            loan_number_norm.isNotNull(),
            F.concat(
                F.lit("LN:"),
                loan_number_norm,
                F.lit("|"),
                lender_for_key,
                F.lit("|"),
                addr_for_key,
            ),
        )
        .when(
            loan_key.isNotNull() & (F.length(F.trim(loan_key)) > 0),
            F.concat(F.lit("NAME:"), loan_key, F.lit("|"), addr_for_key),
        )
        .otherwise(F.lit(None).cast("string"))
    )

    policy_type_upper = F.upper(
        F.trim(F.coalesce(F.col("policy_type_raw"), F.lit("")))
    )
    flood_zone_upper = F.upper(F.trim(F.coalesce(F.col("flood_zone"), F.lit(""))))
    flood_required_inferred_code = (
        F.when(flood_required_code == "REQ", F.lit("REQ"))
        .when(flood_required_code == "NOT_REQ", F.lit("NOT_REQ"))
        .when(
            policy_type_upper.contains("FLOOD") | (policy_type_upper == "NFIP"),
            F.lit("REQ"),
        )
        .when(
            flood_zone_upper.startswith("A") | flood_zone_upper.startswith("V"),
            F.lit("REQ"),
        )
        .when(
            flood_zone_upper.startswith("X")
            | (flood_zone_upper == "B")
            | (flood_zone_upper == "C")
            | (flood_zone_upper == "B/X"),
            F.lit("NOT_REQ"),
        )
        .otherwise(F.lit("UNK"))
    )
    flood_required_inferred_label = (
        F.when(flood_required_inferred_code == "REQ", F.lit("Flood Required"))
        .when(flood_required_inferred_code == "NOT_REQ", F.lit("No Flood Required"))
        .otherwise(F.lit("Unknown"))
    )
    flood_required_inference_source = (
        F.when(
            flood_required_code.isin("REQ", "NOT_REQ"),
            F.lit("raw"),
        )
        .when(
            policy_type_upper.contains("FLOOD") | (policy_type_upper == "NFIP"),
            F.lit("policy-type-inferred"),
        )
        .when(
            flood_zone_upper.startswith("A")
            | flood_zone_upper.startswith("V")
            | flood_zone_upper.startswith("X")
            | (flood_zone_upper == "B")
            | (flood_zone_upper == "C")
            | (flood_zone_upper == "B/X"),
            F.lit("zone-inferred"),
        )
        .otherwise(F.lit("unresolved"))
    )

    # qLenderAliasCurated join not available in this job — fall back to LenderName_Norm.
    lender_canonical = lender_name_norm
    lender_group = lender_name_norm
    lender_alias_is_mapped = F.lit(False)

    fact = combined.select(
        loan_key.alias("loan_key"),
        F.col("source_type"),
        status_normalized.alias("status_normalized"),
        F.col("policy_type_raw"),
        F.col("policy_type_secondary_raw"),
        F.col("flood_required_raw"),
        flood_required_flag.alias("flood_required_flag"),
        is_flood_policy.alias("is_flood_policy"),
        is_hazard_policy.alias("is_hazard_policy"),
        F.col("provider"),
        F.col("lender_name"),
        F.col("carrier"),
        F.col("broker"),
        F.col("state"),
        property_condition.alias("property_condition"),
        year_built.alias("year_built"),
        age_bucket.alias("age_bucket"),
        F.col("flood_zone"),
        _parse_money(F.col("dwelling_amount")).alias("dwelling_amount"),
        _parse_money(F.col("premium_raw")).alias("premium"),
        request_date.alias("request_date"),
        sent_to_procurement_date.alias("sent_to_procurement_date"),
        cert_received_date.alias("cert_received_date"),
        request_month_start.alias("request_month_start"),
        close_bind_date.alias("close_bind_date"),
        close_month_start.alias("close_month_start"),
        closed_flag.alias("closed_flag"),
        days_to_close.alias("days_to_close"),
        F.col("home_buyers_name").alias("home_buyers_name"),
        borrower_name.alias("borrower_name"),
        F.col("arrive_home_loan_number"),
        arrive_home_loan_number_norm.alias("arrive_home_loan_number_norm"),
        F.col("lender_loan_number"),
        F.col("certificate_policy_number"),
        F.col("address"),
        F.col("city"),
        F.col("zip"),
        F.col("county"),
        F.col("processor"),
        F.col("insurance_rep"),
        F.col("status_raw"),
        F.col("approval_status_raw").alias("approval_status"),
        F.col("sqft"),
        effective_date.alias("effective_date"),
        loan_number_norm.alias("loan_number_norm"),
        policy_number_norm.alias("policy_number_norm"),
        loan_key_v2.alias("loan_key_v2"),
        flood_required_code.alias("flood_required_code"),
        flood_required_label.alias("flood_required_label"),
        status_normalized_v2.alias("status_normalized_v2"),
        status_mapped.alias("status_mapped"),
        lender_name_norm.alias("lender_name_norm"),
        address_norm.alias("address_norm"),
        zip_norm.alias("zip_norm"),
        address_key.alias("address_key"),
        loan_key_v3.alias("loan_key_v3"),
        flood_required_inferred_code.alias("flood_required_inferred_code"),
        flood_required_inferred_label.alias("flood_required_inferred_label"),
        flood_required_inference_source.alias("flood_required_inference_source"),
        lender_canonical.alias("lender_canonical"),
        lender_group.alias("lender_group"),
        lender_alias_is_mapped.alias("lender_alias_is_mapped"),
        follow_up_date.alias("follow_up_date"),
        determination_date.alias("determination_date"),
    )

    return fact.filter(
        F.col("loan_key").isNotNull() & (F.length(F.trim(F.col("loan_key"))) > 0)
    )


def _build_dim_loan_master(fact_df: DataFrame) -> DataFrame:
    """One row per Loan_Key_v3 — mirrors the Power Query dim_Loan_Master logic."""
    keyed = fact_df.filter(
        F.col("loan_key_v3").isNotNull() & (F.length(F.col("loan_key_v3")) > 0)
    )
    st_upper = F.upper(F.coalesce(F.col("source_type"), F.lit("")))

    grouped = keyed.groupBy("loan_key_v3").agg(
        F.first("loan_key", ignorenulls=True).alias("loan_key"),
        F.first("loan_key_v2", ignorenulls=True).alias("loan_key_v2"),
        F.first("loan_number_norm", ignorenulls=True).alias("loan_number_norm"),
        F.first("policy_number_norm", ignorenulls=True).alias("policy_number_norm"),
        F.first("lender_name_norm", ignorenulls=True).alias("lender_name_norm"),
        F.first("lender_canonical", ignorenulls=True).alias("lender_canonical"),
        F.first("lender_group", ignorenulls=True).alias("lender_group"),
        F.first("address_norm", ignorenulls=True).alias("address_norm"),
        F.first("zip_norm", ignorenulls=True).alias("zip_norm"),
        F.first("address_key", ignorenulls=True).alias("address_key"),
        F.first("address", ignorenulls=True).alias("address"),
        F.first("city", ignorenulls=True).alias("city"),
        F.first("state", ignorenulls=True).alias("state"),
        F.first("zip", ignorenulls=True).alias("zip"),
        F.first("county", ignorenulls=True).alias("county"),
        F.max(
            F.when(
                st_upper.startswith("INTERNAL") | st_upper.startswith("UNIFICATION"),
                1,
            ).otherwise(0)
        ).alias("has_internal_procured"),
        F.max(F.when(st_upper.startswith("EXTERNAL"), 1).otherwise(0)).alias(
            "has_external"
        ),
        F.count(F.lit(1)).alias("procurement_row_count"),
        F.min("request_date").alias("first_request_date"),
        F.max("request_date").alias("latest_request_date"),
        F.min("close_bind_date").alias("first_close_bind_date"),
        F.max("close_bind_date").alias("latest_close_bind_date"),
        F.max("closed_flag").alias("is_closed_any"),
        F.countDistinct("arrive_home_loan_number_norm").alias(
            "distinct_ah_loan_number_count"
        ),
        F.max(
            F.when(
                F.col("arrive_home_loan_number_norm").isNotNull()
                & (F.length(F.col("arrive_home_loan_number_norm")) > 0),
                F.col("arrive_home_loan_number_norm"),
            )
        ).alias("arrive_home_loan_number_norm"),
        F.collect_set("source_type").alias("_source_types"),
    )

    grouped = grouped.withColumn(
        "has_multiple_source_types",
        F.size(F.col("_source_types")) > 1,
    )

    diag = keyed.groupBy("loan_number_norm").agg(
        F.countDistinct("address_key").alias("distinct_address_key_count"),
        F.countDistinct("lender_name_norm").alias("distinct_lender_count"),
    )

    result = grouped.join(diag, on="loan_number_norm", how="left")

    full_key = (
        F.col("loan_number_norm").isNotNull()
        & (F.length(F.col("loan_number_norm")) > 0)
        & F.col("policy_number_norm").isNotNull()
        & (F.length(F.col("policy_number_norm")) > 0)
        & F.col("lender_name_norm").isNotNull()
        & (F.length(F.col("lender_name_norm")) > 0)
        & F.col("address_key").isNotNull()
        & (F.length(F.col("address_key")) > 0)
    )

    return (
        result.withColumn(
            "key_quality_flag",
            F.when(full_key, F.lit("Full")).otherwise(F.lit("Partial")),
        )
        .withColumn(
            "has_address_conflict",
            F.coalesce(F.col("distinct_address_key_count") > 1, F.lit(False)),
        )
        .withColumn(
            "has_lender_conflict",
            F.coalesce(F.col("distinct_lender_count") > 1, F.lit(False)),
        )
        .drop("_source_types")
        .select(
            "loan_key_v3",
            "loan_key_v2",
            "loan_key",
            "loan_number_norm",
            "arrive_home_loan_number_norm",
            "policy_number_norm",
            "lender_name_norm",
            "lender_canonical",
            "lender_group",
            "address_norm",
            "zip_norm",
            "address_key",
            "address",
            "city",
            "state",
            "zip",
            "county",
            "has_internal_procured",
            "has_external",
            "procurement_row_count",
            "first_request_date",
            "latest_request_date",
            "first_close_bind_date",
            "latest_close_bind_date",
            "is_closed_any",
            "key_quality_flag",
            "distinct_address_key_count",
            "distinct_lender_count",
            "distinct_ah_loan_number_count",
            "has_address_conflict",
            "has_lender_conflict",
            "has_multiple_source_types",
        )
    )


def _build_dim_month(fact_df: DataFrame) -> DataFrame:
    request_months = (
        fact_df.filter(F.col("request_month_start").isNotNull())
        .select(F.col("request_month_start").alias("month_start"))
    )
    close_months = (
        fact_df.filter(F.col("close_month_start").isNotNull())
        .select(F.col("close_month_start").alias("month_start"))
    )
    all_months = request_months.union(close_months).distinct()
    return all_months.select(
        F.col("month_start"),
        F.date_format("month_start", "MMM yyyy").alias("month_label"),
        F.year("month_start").alias("year"),
        F.month("month_start").alias("month_num"),
        (F.year("month_start") * 100 + F.month("month_start")).alias(
            "year_month_sort"
        ),
    )


def _build_dim_date(spark) -> DataFrame:
    start = date(2024, 1, 1)
    end = date(2027, 12, 31)
    cal = spark.sql(
        f"""
        SELECT explode(sequence(
            to_date('{start.isoformat()}'),
            to_date('{end.isoformat()}'),
            interval 1 day
        )) AS date_key
        """
    )
    day_of_week = ((F.dayofweek("date_key") + 5) % 7) + 1
    month_start = F.trunc("date_key", "month")
    return cal.select(
        F.col("date_key").alias("date"),
        F.year("date_key").alias("year"),
        F.month("date_key").alias("month_number"),
        F.date_format("date_key", "MMMM").alias("month_name"),
        F.date_format("date_key", "MMM").alias("month_short"),
        F.date_format("date_key", "yyyy-MM").alias("year_month"),
        (F.year("date_key") * 100 + F.month("date_key")).alias("year_month_sort"),
        F.concat(F.lit("Q"), F.quarter("date_key").cast("string")).alias("quarter"),
        F.concat(
            F.year("date_key").cast("string"),
            F.lit("-Q"),
            F.quarter("date_key").cast("string"),
        ).alias("year_quarter"),
        F.dayofmonth("date_key").alias("day"),
        day_of_week.alias("day_of_week"),
        F.date_format("date_key", "EEEE").alias("day_name"),
        month_start.alias("month_start"),
        F.last_day("date_key").alias("month_end"),
        F.when(day_of_week >= 6, 1).otherwise(0).alias("is_weekend"),
    )


def _write_and_register_table(
    df: DataFrame,
    table_name: str,
    glue_database: str,
    output_s3_base: str,
    write_mode: str,
    dt: str,
    loaded_at: str,
    source_s3_uri: str,
) -> None:
    out_df = (
        df.withColumn("dt", F.lit(dt))
        .withColumn("_etl_loaded_at", F.lit(loaded_at))
        .withColumn("_source_s3_uri", F.lit(source_s3_uri))
    )

    table_location = output_s3_base.rstrip("/") + "/" + table_name + "/"

    if write_mode == "snapshot":
        print(f"  Snapshot mode: clearing previous contents at {table_location}")
        deleted = _delete_s3_prefix(table_location)
        print(f"  Deleted {deleted} previous object(s).")
        out_path = table_location
    else:
        out_path = table_location + f"dt={dt}/"

    print(f"  Writing curated data to: {out_path}")
    (
        out_df.coalesce(1)
        .write.mode("overwrite")
        .option("compression", "snappy")
        .parquet(out_path)
    )
    print("  Write complete.")

    final_schema_cols = [
        (f.name, _spark_type_to_hive(f.dataType))
        for f in out_df.schema.fields
        if not (write_mode == "daily_partition" and f.name == "dt")
    ]
    print(f"  Registering Glue table: {glue_database}.{table_name} (mode={write_mode})")
    _register_glue_table(
        database=glue_database,
        table=table_name,
        location=table_location,
        schema_columns=final_schema_cols,
        partitioned=(write_mode == "daily_partition"),
    )
    print("  Catalog registration complete.")


# =============================================================================
# Main job
# =============================================================================

DEFAULT_SHEET_TABLE_MAP = {
    "Lender procured": "insurance_tracking_lender_procured",
    "Unification of AH procured prem": "insurance_tracking_ah_procured",
}


args = getResolvedOptions(sys.argv, ["JOB_NAME"])

SOURCE_S3_URI = _arg_or_default(
    "--SOURCE_S3_URI",
    "s3://arrive-home-insurance-tracking/InsuranceTrackingSharepoint/latest/InsuranceTracking.xlsx",
)
OUTPUT_S3_BASE = _arg_or_default(
    "--OUTPUT_S3_BASE",
    "s3://arrivehome-bi-prod/curated/",
)
SHEET_TABLE_MAP_JSON = _arg_or_default("--SHEET_TABLE_MAP", "")
GLUE_DATABASE = _arg_or_default("--GLUE_DATABASE", "insurance_tracking_sharepoint")
DT = _arg_or_default("--DT", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
# WRITE_MODE controls how repeated runs behave:
#   snapshot         (default) -> each run REPLACES the table data (no duplicates).
#                                 dt is kept as a regular column for traceability.
#   daily_partition           -> each run writes to dt=YYYY-MM-DD/ (history kept).
#                                 Athena queries must filter by dt to avoid dupes.
WRITE_MODE = _arg_or_default("--WRITE_MODE", "snapshot").strip().lower()
if WRITE_MODE not in ("snapshot", "daily_partition"):
    raise RuntimeError(
        f"Invalid --WRITE_MODE: {WRITE_MODE!r}. Expected 'snapshot' or 'daily_partition'."
    )

if SHEET_TABLE_MAP_JSON.strip():
    try:
        SHEET_TABLE_MAP = json.loads(SHEET_TABLE_MAP_JSON)
        if not isinstance(SHEET_TABLE_MAP, dict):
            raise ValueError("SHEET_TABLE_MAP must be a JSON object")
    except Exception as e:
        raise RuntimeError(
            f"Invalid --SHEET_TABLE_MAP JSON: {SHEET_TABLE_MAP_JSON!r} ({e})"
        )
else:
    SHEET_TABLE_MAP = DEFAULT_SHEET_TABLE_MAP


sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args["JOB_NAME"], args)

spark.conf.set("spark.sql.parquet.datetimeRebaseModeInWrite", "CORRECTED")
spark.conf.set("spark.sql.parquet.int96RebaseModeInWrite", "CORRECTED")

print("=" * 80)
print("INSURANCE TRACKING XLSX → CURATED PARQUET (multi-sheet)")
print("=" * 80)
print(f"Timestamp (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}")
print(f"DT partition: {DT}")
print(f"Source: {SOURCE_S3_URI}")
print(f"Output base: {OUTPUT_S3_BASE}")
print(f"Glue DB: {GLUE_DATABASE}")
print(f"Write mode: {WRITE_MODE}")
print(f"Sheet -> Table map: {json.dumps(SHEET_TABLE_MAP, indent=2)}")

# Download xlsx once and open the workbook
print("\nDownloading XLSX...")
xlsx_bytes = _load_xlsx_bytes_from_s3(SOURCE_S3_URI)
print(f"Downloaded {len(xlsx_bytes):,} bytes")

print("Opening workbook...")
workbook = _open_workbook(xlsx_bytes)
print(f"Sheet names in workbook: {workbook.sheetnames}")

loaded_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
loaded_sheet_dfs: Dict[str, DataFrame] = {}

# Process each requested sheet
for sheet_name, table_name in SHEET_TABLE_MAP.items():
    print("\n" + "-" * 80)
    print(f"PROCESSING sheet '{sheet_name}' -> table '{GLUE_DATABASE}.{table_name}'")
    print("-" * 80)

    if sheet_name not in workbook.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not present; available: {workbook.sheetnames}. Skipping.")
        continue

    raw_rows = _read_sheet_rows(workbook, sheet_name)
    print(f"  Total rows in sheet (incl. header/blank): {len(raw_rows):,}")

    headers, data_rows = _extract_header_and_data(raw_rows)
    print(f"  Original column headers: {headers}")
    print(f"  Data rows after header detection: {len(data_rows):,}")

    if not data_rows:
        print(f"  WARNING: no data rows for sheet '{sheet_name}'. Skipping.")
        continue

    df = _build_spark_df(spark, headers, data_rows)

    # Drop entirely empty/null columns in ONE Spark pass.
    # (Per-column filter().count() in a loop spawns N Spark jobs and is slow on wide sheets.)
    if df.columns:
        agg_exprs = [
            F.sum(
                F.when(
                    F.col(c).isNotNull() & (F.length(F.trim(F.col(c))) > 0), 1
                ).otherwise(0)
            ).alias(c)
            for c in df.columns
        ]
        non_empty_counts = df.agg(*agg_exprs).collect()[0].asDict()
        null_or_empty_cols = [c for c, n in non_empty_counts.items() if (n or 0) == 0]
        if null_or_empty_cols:
            print(f"  Dropping {len(null_or_empty_cols)} fully-empty column(s): {null_or_empty_cols}")
            df = df.drop(*null_or_empty_cols)

    # Normalize column names to snake_case + dedupe
    col_map = _dedupe_columns(df.columns)
    print("  Column rename map (original -> normalized):")
    for old, new in col_map.items():
        print(f"    {old!r:60s} -> {new}")
    for old, new in col_map.items():
        if old != new:
            df = df.withColumnRenamed(old, new)

    # Add partition + metadata columns
    df = (
        df.withColumn("dt", F.lit(DT))
        .withColumn("_etl_loaded_at", F.lit(loaded_at))
        .withColumn("_source_s3_uri", F.lit(SOURCE_S3_URI))
        .withColumn("_source_sheet", F.lit(sheet_name))
    )

    loaded_sheet_dfs[table_name] = df.drop(
        "dt", "_etl_loaded_at", "_source_s3_uri", "_source_sheet"
    )

    _write_and_register_table(
        df=df,
        table_name=table_name,
        glue_database=GLUE_DATABASE,
        output_s3_base=OUTPUT_S3_BASE,
        write_mode=WRITE_MODE,
        dt=DT,
        loaded_at=loaded_at,
        source_s3_uri=SOURCE_S3_URI,
    )

print("\n" + "=" * 80)
print("BUILDING DERIVED TABLES (fact_procurement + dimensions)")
print("=" * 80)

lender_df = loaded_sheet_dfs.get("insurance_tracking_lender_procured")
ah_df = loaded_sheet_dfs.get("insurance_tracking_ah_procured")
fact_procurement_df = _build_fact_procurement(lender_df, ah_df)

if fact_procurement_df is None:
    print("  WARNING: no source sheets loaded; skipping derived tables.")
else:
    derived_tables = [
        ("fact_procurement", fact_procurement_df),
        ("dim_loan_master", _build_dim_loan_master(fact_procurement_df)),
        ("dim_month", _build_dim_month(fact_procurement_df)),
        ("dim_date", _build_dim_date(spark)),
    ]

    for table_name, table_df in derived_tables:
        print("\n" + "-" * 80)
        print(f"PROCESSING derived table '{GLUE_DATABASE}.{table_name}'")
        print("-" * 80)
        _write_and_register_table(
            df=table_df,
            table_name=table_name,
            glue_database=GLUE_DATABASE,
            output_s3_base=OUTPUT_S3_BASE,
            write_mode=WRITE_MODE,
            dt=DT,
            loaded_at=loaded_at,
            source_s3_uri=SOURCE_S3_URI,
        )

print("\nAll sheets processed.")
job.commit()
